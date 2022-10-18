import pandas as pd
from openpyxl import load_workbook
from database import clients

# empiezo leyendo el excel general con este comando
excel = load_workbook('database.xlsx', data_only=True)
# eligo la hoja del archivo de excel que quiero usar
sheet_excel = excel.active
dic = {}
# capturo del excel toda la columna de capital de operacion
capital_operacion = sheet_excel['G2': 'G31']
# capturo del excel toda la columna de clasificacion de dedor
clasificacionDeudor = sheet_excel['K2': 'K31']
# capturo del excel toda la columna de tipo de garantias
tipoGarantia = sheet_excel['I2': 'I31']
# capturo del excel toda la columna de tipo de cartera
tipoCartera = sheet_excel['M2': 'M31']

# recorro la columna de capital de operacion hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
lst_capitalOperacion = []
for fila in capital_operacion:
    for celda in fila:
        lst_capitalOperacion.append(celda.value)
# recorro la columna de capital de clasificacion deudor hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
lst_clasificacionDeudor = []
for fila in clasificacionDeudor:
    for celda in fila:
        lst_clasificacionDeudor.append(celda.value)
# recorro la columna de tipo de garantia hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
lst_tipoGarantia = []
for fila in tipoGarantia:
    for celda in fila:
        lst_tipoGarantia.append(celda.value)
# recorro la columna de tipo de cartera hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
lst_tipoCartera = []
for fila in tipoCartera:
    for celda in fila:
        lst_tipoCartera.append(celda.value)


# creo una lista para guardar los diccionarios los cuales van a contener la informacion obtenida de las listas para cada cliente
lst_dicAux = []  # creacion de lista vacia
for x in range(clients):
    # creacion de los diccionarios vacios
    lst_dicAux.append({'tipoCartera': 0, 'situacionDeudor': 0,
                      'tipoGarantia': 0, 'capital': 0})

aux = 0
aux2 = 0
aux3 = 0
aux4 = 0
# recorro los dicconarios vacios
for dic in lst_dicAux:
    for keys in dic:  # recorro las keys de dicconario por dicconario
        if keys == 'tipoCartera':  # busco asignar los valores a la key tipoCartera
            # asigno valor apartir de las listas de las columnas del excel
            dic[keys] = lst_tipoCartera[aux]
            aux += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
        if keys == 'situacionDeudor':  # busco asignar los valores a la key situacionDeudor
            # asigno valor apartir de las listas de las columnas del excel
            dic[keys] = lst_clasificacionDeudor[aux2]
            aux2 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
        if keys == 'tipoGarantia':  # busco asignar los valores a la key tipoGarantia
            # asigno valor apartir de las listas de las columnas del excel
            dic[keys] = lst_tipoGarantia[aux3]
            aux3 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
        if keys == 'capital':  # busco asignar los valores a la key capital
            # asigno valor apartir de las listas de las columnas del excel
            dic[keys] = lst_capitalOperacion[aux4]
            aux4 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel


# una vez agrupado todos los datos obtenidos en la lista de dicconario, empiezo a filtrar, es decir, utilizar solo los que tienen tipoCartera=1 ya que esta salida requiere solo esos
# lista donde voy a guardar los diccionarios que tienen tipo de cartera 1
lst_carteraConsumo = []
for dic in lst_dicAux:  # recorro los diccionarios de la lista lst_dicAux
    for keys in dic:
        if keys == 'tipoCartera':  # busco analizar la key tipoCartera
            if dic['tipoCartera'] == 1:  # busco solo los que tegna tipoCartera=1
                # guardo los dicconarios de tipoCartera=1
                lst_carteraConsumo.append(dic)


# Creo las siguientes lista para guardar los diccionarios de acuerdo a la situacionDeudor
lst_deudor1 = []
lst_deudor2 = []
lst_deudor3 = []
lst_deudor4 = []
lst_deudor5 = []


for dic in lst_carteraConsumo:  # recorro los dicconarios de la lista de cartera de consumos creada antes
    for keys in dic:  # recorro las keys de diccionario por diccionario
        if keys == 'situacionDeudor':  # analizo la key situacionDeudor
            if dic['situacionDeudor'] == 1:  # busco los que tienen situacionDeudor=1
                # apendeo los diccionarios en la lista correspondiente
                lst_deudor1.append(dic)
            if dic['situacionDeudor'] == 2:  # busco los que tienen situacionDeudor=2
                # apendeo los diccionarios en la lista correspondiente
                lst_deudor2.append(dic)
            if dic['situacionDeudor'] == 3:  # busco los que tienen situacionDeudor=3
                # apendeo los diccionarios en la lista correspondiente
                lst_deudor3.append(dic)
            if dic['situacionDeudor'] == 4:  # busco los que tienen situacionDeudor=4
                # apendeo los diccionarios en la lista correspondiente
                lst_deudor4.append(dic)


'''------------------------ DEUDA 1 ------------------------ '''

# Diccionarios para organizaar la situacion de deduda en este caso 1, de acuerdo al tipo de garantias que tenga
dic_sd1g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
dic_sd1g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
dic_sd1g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

suma_capital0_deuda1 = 0
suma_capital1_deuda1 = 0
suma_capital2_deuda1 = 0

# recorra la lista de los de situacion de deuda 1
for dic in lst_deudor1:
    for keys in dic:
        if keys == 'tipoGarantia':
            if dic['tipoGarantia'] == 0:
                # sumo el capital de todos aquellos que no tengan garantia (0)
                suma_capital0_deuda1 += dic['capital']
            if dic['tipoGarantia'] == 1:
                # sumo el capital de todos aquellos que tengan garantia A (1)
                suma_capital1_deuda1 += dic['capital']
            if dic['tipoGarantia'] == 2:
                # sumo el capital de todos aquellos que tengan garantia B (2)
                suma_capital2_deuda1 += dic['capital']


# luego organizo todos los diccionarios con la suma de capital obtenida arriba
dic_sd1g0['capitalTotal'] = suma_capital0_deuda1
dic_sd1g1['capitalTotal'] = suma_capital1_deuda1
dic_sd1g2['capitalTotal'] = suma_capital2_deuda1

# re-organizo la lista de deudor con los diccionarios de arriba
lst_deudor1 = [0, 0, 0]
lst_deudor1[0] = dic_sd1g0
lst_deudor1[1] = dic_sd1g1
lst_deudor1[2] = dic_sd1g2

# print(lst_deudor1)


'''------------------------ DEUDA 2 ------------------------ '''

# Diccionarios para organizaar la situacion de deduda en este caso 2, de acuerdo al tipo de garantias que tenga
dic_sd2g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
dic_sd2g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
dic_sd2g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

suma_capital0_deuda2 = 0
suma_capital1_deuda2 = 0
suma_capital2_deuda2 = 0

# recorra la lista de los de situacion de deuda 2
for dic in lst_deudor2:
    for keys in dic:
        if keys == 'tipoGarantia':
            if dic['tipoGarantia'] == 0:
                # sumo el capital de todos aquellos que no tengan garantia (0)
                suma_capital0_deuda2 += dic['capital']
            if dic['tipoGarantia'] == 1:
                # sumo el capital de todos aquellos que tengan garantia A (1)
                suma_capital1_deuda2 += dic['capital']
            if dic['tipoGarantia'] == 2:
                # sumo el capital de todos aquellos que tengan garantia B (2)
                suma_capital2_deuda2 += dic['capital']


# luego organizo todos los diccionarios con la suma de capital obtenida arriba
dic_sd2g0['capitalTotal'] = suma_capital0_deuda2
dic_sd2g1['capitalTotal'] = suma_capital1_deuda2
dic_sd2g2['capitalTotal'] = suma_capital2_deuda2

# re-organizo la lista de deudor con los diccionarios de arriba
lst_deudor2 = [0, 0, 0]
lst_deudor2[0] = dic_sd2g0
lst_deudor2[1] = dic_sd2g1
lst_deudor2[2] = dic_sd2g2

# print(lst_deudor2)

'''------------------------ DEUDA 3 ------------------------ '''

# Diccionarios para organizaar la situacion de deduda en este caso 3, de acuerdo al tipo de garantias que tenga
dic_sd3g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
dic_sd3g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
dic_sd3g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

suma_capital0_deuda3 = 0
suma_capital1_deuda3 = 0
suma_capital2_deuda3 = 0

# recorra la lista de los de situacion de deuda 3
for dic in lst_deudor3:
    for keys in dic:
        if keys == 'tipoGarantia':
            if dic['tipoGarantia'] == 0:
                # sumo el capital de todos aquellos que no tengan garantia (0)
                suma_capital0_deuda3 += dic['capital']
            if dic['tipoGarantia'] == 1:
                # sumo el capital de todos aquellos que tengan garantia A (1)
                suma_capital1_deuda3 += dic['capital']
            if dic['tipoGarantia'] == 2:
                # sumo el capital de todos aquellos que tengan garantia B (2)
                suma_capital2_deuda3 += dic['capital']


# luego organizo todos los diccionarios con la suma de capital obtenida arriba
dic_sd3g0['capitalTotal'] = suma_capital0_deuda3
dic_sd3g1['capitalTotal'] = suma_capital1_deuda3
dic_sd3g2['capitalTotal'] = suma_capital2_deuda3

# re-organizo la lista de deudor con los diccionarios de arriba
lst_deudor3 = [0, 0, 0]
lst_deudor3[0] = dic_sd3g0
lst_deudor3[1] = dic_sd3g1
lst_deudor3[2] = dic_sd3g2

# print(lst_deudor3)
'''------------------------ DEUDA 4 ------------------------ '''

# creo diccionarios para organizar la situacion de deuda en este caso 4, de acuerdo al tipo de garantia que tenga
dic_sd4g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
dic_sd4g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
dic_sd4g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

suma_capital0_deuda4 = 0
suma_capital1_deuda4 = 0
suma_capital2_deuda4 = 0

# recorra la lista de los de situacion de deuda 4
for dic in lst_deudor4:
    for keys in dic:
        if keys == 'tipoGarantia':
            if dic['tipoGarantia'] == 0:
                # sumo el capital de todos aquellos que no tengan garantia (0)
                suma_capital0_deuda4 += dic['capital']
            if dic['tipoGarantia'] == 1:
                # sumo el capital de todos aquellos que tengan garantia A (1)
                suma_capital1_deuda4 += dic['capital']
            if dic['tipoGarantia'] == 2:
                # sumo el capital de todos aquellos que tengan garantia B (2)
                suma_capital2_deuda4 += dic['capital']

# luego organizo todos los diccionarios con la suma de capital obtenida arriba
dic_sd4g0['capitalTotal'] = suma_capital0_deuda4
dic_sd4g1['capitalTotal'] = suma_capital1_deuda4
dic_sd4g2['capitalTotal'] = suma_capital2_deuda4

# re-organizo la lista de deudor con los diccionarios de arriba
lst_deudor4 = [0, 0, 0]
lst_deudor4[0] = dic_sd4g0
lst_deudor4[1] = dic_sd4g1
lst_deudor4[2] = dic_sd4g2
# print(lst_deudor4)


'''------------------------ DEUDA 5 ------------------------ '''

# creo diccionarios para organizar la situacion de deuda en este caso 5, de acuerdo al tipo de garantia que tenga
dic_sd5g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
dic_sd5g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
dic_sd5g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

suma_capital0_deuda5 = 0
suma_capital1_deuda5 = 0
suma_capital2_deuda5 = 0

# recorra la lista de los de situacion de deuda 5
for dic in lst_deudor5:
    for keys in dic:
        if keys == 'tipoGarantia':
            if dic['tipoGarantia'] == 0:
                # sumo el capital de todos aquellos que no tengan garantia (0)
                suma_capital0_deuda5 += dic['capital']
            if dic['tipoGarantia'] == 1:
                # sumo el capital de todos aquellos que tengan garantia A (1)
                suma_capital1_deuda5 += dic['capital']
            if dic['tipoGarantia'] == 2:
                # sumo el capital de todos aquellos que tengan garantia B (2)
                suma_capital2_deuda5 += dic['capital']

# luego organizo todos los diccionarios con la suma de capital obtenida arriba
dic_sd5g0['capitalTotal'] = suma_capital0_deuda5
dic_sd5g1['capitalTotal'] = suma_capital1_deuda5
dic_sd5g2['capitalTotal'] = suma_capital2_deuda5

# re-organizo la lista de deudor con los diccionarios de arriba
lst_deudor5 = [0, 0, 0]
lst_deudor5[0] = dic_sd5g0
lst_deudor5[1] = dic_sd5g1
lst_deudor5[2] = dic_sd5g2
# print(lst_deudor5)


# lista data frame
lista_deudores = []


for x in lst_deudor1:
    lista_deudores.append(x)

for x in lst_deudor2:
    lista_deudores.append(x)

for x in lst_deudor3:
    lista_deudores.append(x)

for x in lst_deudor4:
    lista_deudores.append(x)

for x in lst_deudor5:
    lista_deudores.append(x)
# lista_deudores.append(lst_deudor1)
# lista_deudores.append(lst_deudor2)
# lista_deudores.append(lst_deudor3)
# lista_deudores.append(lst_deudor4)
# lista_deudores.append(lst_deudor5)


print(lista_deudores, 'lista deudores')

# Lista para generar el excel
datosDeudores = pd.DataFrame(lista_deudores)

# creacion del excel
datosDeudores.to_excel('datos_deudores.xlsx')


# ruta de nuestro archivo
filesheet = "./datos_deudores.xlsx"

# creamos el objeto load_workbook
wb = load_workbook(filesheet)

# Seleccionamos el archivo
sheet = wb.active

# Deudor 1 - COlumna excel

sheet['A1'] = 'Situacion deudor'
sheet['A2'] = '1'
sheet['A3'] = '1'
sheet['A4'] = '1'

sheet['A5'] = '2'
sheet['A6'] = '2'
sheet['A7'] = '2'

sheet['A8'] = '3'
sheet['A9'] = '3'
sheet['A10'] = '3'

sheet['A11'] = '4'
sheet['A12'] = '4'
sheet['A13'] = '4'


sheet['A14'] = '5'
sheet['A15'] = '5'
sheet['A16'] = '5'
# Guardamos el archivo con los cambios
wb.save(filesheet)
